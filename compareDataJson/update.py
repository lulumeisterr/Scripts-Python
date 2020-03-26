import mysql.connector
import json
import logging
import re
from openpyxl import Workbook , load_workbook

#----------------------------------------------------------------------------

# Ler o arquivo JSON e comparar as intents

myFile=open('C:\Users\Lrodrina\Desktop\skillbot.json','r')
myObject=myFile.read()
u = unicode(myObject.decode('utf-8-sig'))
myObject = u.encode('utf-8')
myFile.encoding
myFile.close()
myData=json.loads(myObject,'utf-8')

#----------------------------------------------------------------------------

wb = Workbook()
wb = load_workbook('C:/Users/Lrodrina/Desktop/arq1.xlsx')
ws = wb.active
wb.template = False

#----------------------------------------------------------------------------
def checkJson(jsonContents):
	bodyFlag = True if "output" in myData['dialog_nodes'][0] else False
	return bodyFlag

arrayIntentsComDialog = []
removendoLinhas = []
sheet = wb.get_sheet_by_name('Relatorio Resumo')

i = 0
k = 0
while(i < len(myData['dialog_nodes'])):
	try:
		if(checkJson(json.dumps(myData['dialog_nodes'][i]['output']))):
			if(unicode(json.dumps(myData['intents']) in unicode(json.dumps(myData['dialog_nodes'][i]['output']['generic'][0]['values'])))):
				#print('Deu certo : ' + unicode(myData['dialog_nodes'][i]['output']['generic'][0]['values']))
				arrayIntentsComDialog.append( unicode(myData['dialog_nodes'][i]['output']['generic'][0]['values']))
				j = 0
				while(j < len(arrayIntentsComDialog)):
					ws.cell(column=1, row=1,value='Intents que possui Dialogs')
					ws.cell(column=1, row=j+2,value=arrayIntentsComDialog[j].replace('{','').replace('}','').replace(']','').replace('[','').replace("u'text':","").replace("u'","").replace("'","").strip())	
					j = j + 1

					if(sheet.cell(row = j+1, column=1).value == ''):
						removendoLinhas.append(j)	
					else:
						print("No blank : " + str(sheet.cell(row = j+1, column=1).value))

				wb.save('/Users/Lrodrina/Desktop/nodes.xlsx')
			else:
				print('deu ruim')
	except KeyError as error:
		print('Nao foi possivel localizar esse valor no objeto : ' + str(error))
	except Exception as exception:
		logging.exception(exception)
	i = i + 1


#for r in removendoLinhas:
	#print(r)

#-----------------------------------------------------------------------------





	

