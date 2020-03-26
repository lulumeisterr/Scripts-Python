from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import Workbook
import datetime
import gspread
import mysql.connector
import json
import time
import os
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

#------------------------------------------------------------------------------
			#------------------VARIAVEIS DATA -------------------#
date_now = datetime.datetime.now()

date_afterMidNight = date_now.replace(hour=0, minute=0, second=0)
date_beforeMidNight = date_now.replace(hour=23, minute=59, second=59)

one_days_ago = date_afterMidNight - datetime.timedelta(days=1)
one_days_ago_after_23hours = date_beforeMidNight - datetime.timedelta(days=1)

DT_TEMPLATE = one_days_ago.strftime("%d/%m/%Y")
DT_INI = one_days_ago.strftime("%Y-%m-%d %H:%M:%S")
DT_FIM = one_days_ago_after_23hours.strftime("%Y-%m-%d %H:%M:%S")

print(DT_INI)
print(DT_FIM)
#------------------------------------------------------------------------------

armazenaDadosQuery = {
	'Data' : '',
	'textoUser':'',
	'intencao':'',
	'textElo':'',
	'sessionCode':'',
	'UserRef':'',
}	


def enviarEmail():

	#filepath = "C:/Users/Lrodrina/Desktop/googleShetsRelatorioEmail/RELATORIO_INTERACAO.xlsx"
	filepath = "/exemplo/googleSheetsAPI/sendEmailRelatorio/RELATORIO_INTERACAO.xlsx"
	
	msgFrom = "exeplo.com"
	smtpObj = smtplib.SMTP('smtp.live.com', 587)
	smtpObj.ehlo()
	smtpObj.starttls()
	msgTo = '#'
	toPass = '#'
	smtpObj.login(msgTo, toPass)

	msg = MIMEMultipart()

	attachment = open(filepath,'rb')
	xlsx = MIMEBase('application','vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	xlsx.set_payload(attachment.read())

	encoders.encode_base64(xlsx)
	xlsx.add_header('Content-Disposition', 'attachment', filename=filepath)
	msg.attach(xlsx)

	smtpObj.sendmail(msgTo,msgFrom,msg.as_string())
	smtpObj.quit()
	print("Email enviado com sucesso!")
	


def querysMysql():

	#filepath = "C:/Users/Lrodrina/Desktop/googleShetsRelatorioEmail/RELATORIO_INTERACAO.xlsx"
	filepath = "/root/googleSheetsAPI/sendEmailRelatorio/RELATORIO_INTERACAO.xlsx"

	fieldnames = ['Data' , 'Texto do usuario' , 'Intencao', 'textoElo','sessionCode','userID']

	wb = Workbook()
	wb.save(filepath)

	ws = wb.active
	ws = wb.create_sheet('Relatorio')

	# Configuracoes do banco de dados Hml
	#con = mysql.connector.connect(user='#', password='#',host='#',database='#')

	# Configuracoes do banco de dados Prod
	con = mysql.connector.connect(user='#', password='#',host='#',database='#')
	#----------------------------------------------------------------------------
	# Criando Statemant
	c = con.cursor(buffered=True , dictionary=True)
	#----------------------------------------------------------------------------
	#---------------###################QUERYS########################---------------#
	print("Executando Query")

	queryDiaria = """
		(SELECT DATE_FORMAT(`ui`.`createDate`,'%d-%m-%Y') AS `Data`,`ui`.`text` AS `textoUser`,`a`.`code` AS `intencao`,`a`.`text` AS `textoElo`,`ui`.`sessionCode` AS `sessionCode`,
		se.userRef AS 'userId'
		FROM (`user_interaction` AS ui LEFT JOIN session as se ON se.sessionCode = ui.sessionCode INNER JOIN answer ans ON ans.id = ui.answerId
		JOIN `answer` `a` ON(((`ui`.`answerId` = `a`.`id`))))
		WHERE ans.projectId = 4 AND se.userRef NOT IN (2039613406150750,2357723287650147,3819788814713616,2120042778096110,2546582828765265,2556592324376460,2149525391823920) 
		AND ui.text IS NOT NULL AND a.code NOT IN ('20101_SAUDACAO','20101_SAUDACAO_2') AND (`ui`.`createDate` > '{0}') AND (`ui`.`createDate` < '{1}') AND ui.text NOT IN ('qQs!+eTCh-CDX9a^')
		)
      """
	c.execute(queryDiaria.format(DT_INI,DT_FIM))
	rowsQueryDiaria = c.fetchall()

	coluna = 1
	linha = 2
	v = 0

	print(len(fieldnames))

	j = 1
	while(j < len(fieldnames) + 1):
		ws.cell(column=j, row=1,value=fieldnames[v])
		print(fieldnames[v])
		j = j + 1
		v = v + 1

	for i in rowsQueryDiaria:

		armazenaDadosQuery['Data'] = i['Data']
		armazenaDadosQuery['textoUser'] = i['textoUser']
		armazenaDadosQuery['intencao'] = i['intencao']
		armazenaDadosQuery['textoElo'] = i['textoElo']
		armazenaDadosQuery['sessionCode'] = i['sessionCode']
		armazenaDadosQuery['userRef'] = i['userId']
		
		ws.cell(column=1, row=linha,value=armazenaDadosQuery['Data'])
		ws.cell(column=2, row=linha,value=armazenaDadosQuery['textoUser'])
		ws.cell(column=3, row=linha,value=armazenaDadosQuery['textoElo'])
		ws.cell(column=4, row=linha,value=armazenaDadosQuery['intencao'])
		ws.cell(column=5, row=linha,value=armazenaDadosQuery['sessionCode'])
		ws.cell(column=6, row=linha,value=armazenaDadosQuery['userRef'])

		linha = linha + 1

	wb.remove(wb["Sheet"])
	wb.save(filepath)

	enviarEmail()
    #----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
querysMysql()

